"""
점검보고서 자동 분석 도구
사용법: python analyze_reports.py [폴더경로]
       폴더경로 생략 시 현재 디렉터리 사용

출력: 분석결과_YYYYMMDD.xlsx
"""
import sys
import re
import os
from pathlib import Path
from datetime import datetime

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("필수 패키지 설치 필요: pip install pdfplumber openpyxl")
    sys.exit(1)


# ── 점검 종류 감지 ──────────────────────────────────────────────
def detect_insp_type(fname: str) -> str | None:
    if re.search(r'정기안전점검|3종시설물|3종점검|정기점검', fname):
        return '3종시설물 정기점검'
    if re.search(r'정밀안전진단|정밀진단', fname):
        return '정밀안전진단'
    if re.search(r'정밀안전점검|정밀점검', fname):
        return '정밀안전점검'
    return None


# ── PDF 텍스트 추출 ────────────────────────────────────────────
def extract_pdf_text(path: Path, max_pages: int = 10) -> tuple[str, bool]:
    """(text, is_scanned) 반환"""
    if pdfplumber is None:
        return '', False
    try:
        with pdfplumber.open(path) as pdf:
            text = ''
            pages = min(len(pdf.pages), max_pages)
            for i in range(pages):
                t = pdf.pages[i].extract_text() or ''
                text += t + '\n'
            scanned = len(text.strip()) < 50
            return text, scanned
    except Exception as e:
        return '', False


# ── 연도·분기·등급 파싱 ────────────────────────────────────────
def parse_insp_info(text: str) -> dict:
    year, period, grade = None, None, None

    # 연도 + 월
    for m in re.finditer(r'(\d{4})\s*[.년]\s*0?(\d{1,2})\s*[.월]', text):
        y, mo = int(m.group(1)), int(m.group(2))
        if 2010 <= y <= 2035 and 1 <= mo <= 12:
            year = y
            period = '상반기' if mo <= 6 else '하반기'
            break

    # 안전등급 (우선순위 순)
    grade_patterns = [
        r'종합\s*안전\s*등급\s*[:\s]*([A-Ea-e])',
        r'안전\s*등급\s*[:\s]*([A-Ea-e])',
        r'상태\s*평가\s*[:\s]*([A-Ea-e])',
        r'([A-Ea-e])\s*등급',
    ]
    for pat in grade_patterns:
        m = re.search(pat, text)
        if m:
            grade = m.group(1).upper()
            break

    return {'year': year, 'period': period, 'grade': grade}


# ── 파일 분석 ──────────────────────────────────────────────────
def analyze_file(path: Path) -> dict:
    result = {
        'filename': path.name,
        'path': str(path),
        'insp_type': detect_insp_type(path.name),
        'year': None,
        'period': None,
        'grade': None,
        'status': '',
    }

    if result['insp_type'] is None:
        result['status'] = '점검보고서 아님'
        return result

    suffix = path.suffix.lower()
    if suffix != '.pdf':
        result['status'] = f'PDF 아님 ({suffix}) — 수동확인 필요'
        return result

    text, scanned = extract_pdf_text(path)
    if scanned or not text.strip():
        result['status'] = '스캔PDF — 수동확인 필요'
        return result

    info = parse_insp_info(text)
    result.update(info)

    missing = []
    if not info['year']:   missing.append('연도')
    if not info['grade']:  missing.append('등급')
    result['status'] = '완료' if not missing else f'미감지: {", ".join(missing)}'
    return result


# ── Excel 출력 ─────────────────────────────────────────────────
GRADE_COLORS = {
    'A': 'C6EFCE', 'B': 'DDEBF7', 'C': 'FFEB9C',
    'D': 'FFC7CE', 'E': 'FF0000',
}

def write_excel(results: list[dict], out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '점검보고서 분석결과'

    # 헤더
    headers = ['#', '파일명', '점검종류', '연도', '분기', '안전등급', '상태', '경로']
    ws.append(headers)
    hdr_fill = PatternFill('solid', fgColor='2F5496')
    hdr_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['H'].width = 60

    ok_fill   = PatternFill('solid', fgColor='E2EFDA')
    warn_fill = PatternFill('solid', fgColor='FFF2CC')
    err_fill  = PatternFill('solid', fgColor='FFE0E0')
    scan_fill = PatternFill('solid', fgColor='F2F2F2')

    insp_only = [r for r in results if r['insp_type']]
    other     = [r for r in results if not r['insp_type']]

    for i, r in enumerate(insp_only + other, 1):
        year_str = f"{r['year']}년 {r['period']}" if r['year'] and r['period'] else (str(r['year']) if r['year'] else '')
        row = [i, r['filename'], r['insp_type'] or '', year_str, r['period'] or '', r['grade'] or '', r['status'], r['path']]
        ws.append(row)
        rn = ws.max_row

        # 상태별 행 색
        if r['status'] == '완료':
            fill = ok_fill
        elif '스캔' in r['status']:
            fill = scan_fill
        elif '미감지' in r['status']:
            fill = warn_fill
        elif r['insp_type'] is None:
            fill = None
        else:
            fill = err_fill

        if fill:
            for c in ws[rn]:
                c.fill = fill

        # 등급 셀 색
        grade_cell = ws.cell(rn, 6)
        if r['grade'] and r['grade'] in GRADE_COLORS:
            grade_cell.fill = PatternFill('solid', fgColor=GRADE_COLORS[r['grade']])
            grade_cell.font = Font(bold=True)
        grade_cell.alignment = Alignment(horizontal='center')

    # 요약 통계
    ws.append([])
    total   = len(insp_only)
    done    = sum(1 for r in insp_only if r['status']=='완료')
    scanned = sum(1 for r in insp_only if '스캔' in r['status'])
    no_grd  = sum(1 for r in insp_only if '미감지' in r['status'])
    ws.append(['', '== 요약 ==', f'총 {total}건', f'완료 {done}건', f'스캔PDF {scanned}건', f'등급미감지 {no_grd}건'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    wb.save(out_path)
    print(f"\n✅ 저장 완료: {out_path}")


# ── 메인 ──────────────────────────────────────────────────────
def main():
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('.')
    if not folder.exists():
        print(f"폴더를 찾을 수 없습니다: {folder}")
        sys.exit(1)

    if pdfplumber is None:
        print("⚠️  pdfplumber 미설치 — PDF 텍스트 추출 불가\n   pip install pdfplumber")

    pdfs = list(folder.rglob('*.pdf')) + list(folder.rglob('*.PDF'))
    if not pdfs:
        print("PDF 파일이 없습니다.")
        sys.exit(0)

    print(f"📂 {folder}")
    print(f"   PDF {len(pdfs)}개 발견 — 분석 시작...\n")

    results = []
    for i, p in enumerate(pdfs, 1):
        r = analyze_file(p)
        results.append(r)
        grade_str = f" [{r['grade']}등급]" if r['grade'] else ''
        year_str  = f" {r['year']}년 {r['period']}" if r['year'] else ''
        print(f"  [{i:3}/{len(pdfs)}] {p.name[:45]:<45} → {r['status']}{year_str}{grade_str}")

    out = folder / f"분석결과_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    write_excel(results, out)

    # 수동 확인 필요 목록 출력
    needs_check = [r for r in results if r['status'] not in ('완료', '점검보고서 아님')]
    if needs_check:
        print(f"\n⚠️  수동 확인 필요 ({len(needs_check)}건):")
        for r in needs_check:
            print(f"   • {r['filename']}: {r['status']}")


if __name__ == '__main__':
    main()
